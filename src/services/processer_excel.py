import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import threading

columns_name = [
    "po_no",
    "start_date",
    "end_date",
    "company_name",
    "order_name",
    "amount",
    "unit",
    "price",
    "total",
    "price_vat",
    "total_vat"
]

class ProcessorExcel:
    def __init__(self, history_file_path=None, orders_file_path=None):
        self.history_file_path = history_file_path
        self.orders_file_path = orders_file_path
        self.date = datetime.today() + relativedelta(years=543)
        self.current_row = 0
        self.total_row = 0

    def process(self):
        self.history_df = pd.read_excel(
            self.history_file_path,
            sheet_name="บริษัท, ร้านค้า", 
            usecols=range(0, len(columns_name)))
        self.history_df.columns = columns_name

        self.orders_df = pd.read_excel(
            self.orders_file_path,
            sheet_name="วัตถุดิบ",
            usecols=range(0, len(columns_name)),
        )
        self.orders_df.columns = columns_name

        self.orders_wb = load_workbook(self.orders_file_path)

        self.total_row = self.orders_df.shape[0]

        for index, row_series in self.orders_df.iterrows():
            order = row_series.to_dict()
            order_name = order["order_name"]
            self.current_row += 1

            for _, row_series_history_order in self.history_df.iterrows():
                history_order = row_series_history_order.to_dict()
                history_order_name = history_order["order_name"]
                try:
                    if history_order_name == order_name and history_order["start_date"] <= self.date:
                        self.orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("po_no") + 1).value = history_order["po_no"]
                        self.orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("start_date") + 1).value = history_order["start_date"]
                        self.orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("end_date") + 1).value = history_order["end_date"]
                        self.orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("company_name") + 1).value = history_order["company_name"]
                        self.orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("price") + 1).value = history_order["price"]
                except Exception as e:
                    pass

    def save(self, save_path):
        self.orders_wb.save(save_path)

    def get_progress(self):
        try:
            return (self.current_row / self.total_row) * 100
        except ZeroDivisionError:
            return 0
        
processor = ProcessorExcel()
