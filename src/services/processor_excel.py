import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from dateutil.relativedelta import relativedelta

columns_name = [
    "po_no",
    "start_date",
    "end_date",
    "company_name",
    "order_name",
    "amount",
    "unit",
    "price",
    "total",
    "price_vat",
    "total_vat"
]

class ProcessorExcel:

    def __init__(self, history_file_path=None, orders_file_path=None):
        self.__history_file_path = history_file_path
        self.__orders_file_path = orders_file_path
        self.__date = datetime.today() + relativedelta(years=543)

        self.__history_df = None
        self.__orders_df = None
        self.__orders_wb = None
        self.__total_row = None
        self.__current_row = None

    def run_pre_process(self):
        self.__history_df = pd.read_excel(
            self.__history_file_path,
            sheet_name="บริษัท, ร้านค้า", 
            usecols=range(0, len(columns_name)))
        self.__history_df.columns = columns_name

        self.__orders_df = pd.read_excel(
            self.__orders_file_path,
            sheet_name="วัตถุดิบ",
            usecols=range(0, len(columns_name))
        )
        self.__orders_df.columns = columns_name

        self.__orders_wb = load_workbook(self.__orders_file_path)

        self.__total_row = self.__orders_df.shape[0]
        self.__current_row = 0

    def run_main_process(self):
        self.run_pre_process()

        for index, row_series in self.__orders_df.iterrows():
            self.__current_row += 1
            order = row_series.to_dict()

            try:
                order_name  = order["order_name"].strip()
                order_unit = order["unit"].strip()
            except:
                continue

            if order_name == "" or order_name == None:
                continue

            if order_unit == "" or order_unit == None:
                continue

            if "{" in order_name:
                order_name = re.sub(r"\{.*?\}", "", order_name)
                order_name = re.sub(r"\s+", " ", order_name).strip()

            for _, row_series_history_order in self.__history_df[::-1].iterrows():

                try:
                    history_order = row_series_history_order.to_dict()
                    history_order_name = history_order["order_name"].strip()
                    history_order_unit = history_order["unit"].strip()

                    if not history_order["start_date"] <= self.__date:
                        continue

                    if "{" in history_order_name:
                        history_order_name = re.sub(r"\{.*?\}", "", history_order_name)
                        history_order_name = re.sub(r"\s+", " ", history_order_name).strip()

                    if not order_name == history_order_name:
                        continue

                    if not order_unit == history_order_unit:
                        cell = self.__orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("po_no") + 1)
                        cell.value = "หน่วยไม่ตรงกัน"
                        cell.font = Font(color="980000")
                        break

                    self.__orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("po_no") + 1).value = history_order["po_no"]
                    self.__orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("start_date") + 1).value = history_order["start_date"]
                    self.__orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("end_date") + 1).value = history_order["end_date"]
                    self.__orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("company_name") + 1).value = history_order["company_name"]
                    self.__orders_wb["วัตถุดิบ"].cell(row=index + 2, column=columns_name.index("price") + 1).value = history_order["price"]

                    break

                except:
                    pass

    def save(self, save_path):
        self.__orders_wb.save(save_path)

    def get_progress(self):
        try:
            return (self.__current_row / self.__total_row) * 100
        except Exception:
            return 0
        
    def set_date(self, date):
        self.__date = date
    
    def set_history_file_path(self, history_file_path):
        self.__history_file_path = history_file_path
    
    def set_orders_file_path(self, orders_file_path):
        self.__orders_file_path = orders_file_path
        
processor = ProcessorExcel()
